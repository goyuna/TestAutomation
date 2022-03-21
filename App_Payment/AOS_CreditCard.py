#실행 : python AOS_KR.py
import base64
import unittest
import HtmlTestRunner
import os
import time
import openpyxl   
from appium import webdriver
from time import sleep
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException

class CreditCardTest(unittest.TestCase):

    #테스트 실행 전 수행해야하는 코드
    @classmethod
    def setUpClass(cls):

        #Set up Directory & Workbook
        tm = time.localtime()
        str_tm = time.strftime('%Y-%m-%d %H:%M:%S', tm)

        wb=openpyxl.Workbook()
        sheet=wb.active
        sheet['A1'] = str_tm
        sheet.append(['법인명', '계정명', '주문 금액', '주문 번호', '취소 확인', '비고'])
        wb.save('APP_CreditCard.xlsx')

    #각 TC 실행 전 수행되는 코드
    def setUp(self):
        app = os.path.join(os.path.dirname(__file__), 'C:\\Users\\고유나\\Desktop\\','m.com.atomy_50_apps.evozi.com.apk')
        app = os.path.abspath(app)

        #Set up appium
        self.driver = webdriver.Remote(
            command_executor='http://127.0.0.1:4723/wd/hub',
            desired_capabilities={
                'app': app,
                'platformName': 'Android',
                'platformVersion': '9.0',
                'deviceName': 'GalaxyS8',
                'automationName': 'uiautomator2',
                'appPackage': 'm.com.atomy',
                'appActivity': 'com.atomy.android.app.views.activities.splash.SplashActivity',
                'udid': 'ce021712d2a8d11e05',
                'chromeOptions' : {'w3c' : False},
                #'noReset': True
            })

    #실행할 테스트 케이스
    def test_KR_CreditCard(self):
        KRID = 'tetID'
        PW = 'test'
        Order_Info = {'Nation': 'KR', 'ID': KRID, 'Price': '', 'Num': '', 'Cancel':'', 'Etc': ''}

    #=====>앱 실행 및 권한 허용

        # appium의 webdriver를 초기화
        driver = self.driver

        # selenium의 webdriverwait을 사용. element가 나올 때 까지 최대 20초까지 wait
        wait = WebDriverWait(driver, 20)

        #화면 녹화 시작
        #self.driver.start_recording_screen()    

        # 권한 허용 버튼 클릭
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.ID,'com.android.packageinstaller:id/permission_allow_button')))
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.ID,'com.android.packageinstaller:id/permission_allow_button')))
        el.click()
        sleep(30)

        #=====>웹뷰로 전환
        self.driver.switch_to.context('WEBVIEW_m.com.atomy')

        #공지사항 하루종일 보지않기
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//section[@id="popup"]/div/div[2]/div/span[1]')))
        el.click()
        sleep(2)

    #=====>로그인
        #로그인버튼 클릭
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//header[@id="header"]/div[2]/span[1]/a')))
        el.click()
        sleep(2)

        #로그인창으로 Switch
        window_handles = self.driver.window_handles
        #print(window_handles)
        self.driver.switch_to.window(window_handles[1])

        #로그인 진행
        id = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@id="UserId"]')))
        id.send_keys(KRID)
        pw = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@id="Password"]')))
        pw.send_keys(PW)
        sleep(1)
        self.driver.find_element_by_xpath('//button[@id="btnLogin"]').click()
        sleep(3)

    #=====>쇼핑몰로 이동
        self.driver.switch_to.window(window_handles[0])

        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//nav[@id="nav"]/a[1]')))
        el.click()
        sleep(5)

        #쇼핑몰창으로 Switch
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[2])

        #튜토리얼 닫기
        self.driver.find_element_by_xpath('//section[@id="toutorialDialog"]/div/button').click()
        sleep(2)

    #=====>최저가격 상품 찾기
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//select[@name="orderType"]/option[3]')))     #낮은 가격순
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//ul[@id="list"]/li[2]/div/div/div/span/a')))
        el.click()
        sleep(2)

        #바로구매 클릭
        window_handles = self.driver.window_handles
        #print(window_handles)
        self.driver.switch_to.window(window_handles[3])
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnBuy"]')))
        el.click() 
        sleep(1)
        el.click()
        sleep(1)

    #=====>배송 정보 입력
        sleep(3)
        self.driver.switch_to.window(window_handles[2])
        self.driver.execute_script('window.scrollTo(0, 600)')
        sleep(1)                                     
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//section[@id="tabInsertShippingInfo"]')))
        el.click() 
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//div[@id="radioShippingInfo"]')))   #주문자와 동일
        el.click()

        #결제 방법 입력
        self.driver.execute_script('window.scrollTo(0, 1800)') 
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//li[@id="item2"]'))) #신용카드 - inipay 선택
        el.click() 

    #=====>결제 하기
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        sleep(1)
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//input[@id="checkBuyAgree"]')))  #동의하기 선택
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnBuy"]')))  #결제하기 클릭
        el.click() 
        sleep(10)
        
        #KG이니시스 화면으로 이동
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[5])

        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="termsArea"]/div[2]/section/div[1]/div[2]/div/div')))  #이용약관 전체 동의 클릭
        el.click()
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="payArea"]/div[5]/ul/li[6]/span')))  #더보기 클릭
        el.click()
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="payArea"]/div[5]/ul/li[9]/span')))  #우리은행 선택
        el.click()

        el = wait.until(EC.presence_of_element_located((By.ID, 'cardNext2Btn')))  #다음 클릭
        el.click()

        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="wrap"]/div/table[1]/tbody/tr/td[2]')))  #기타 결제 클릭
        el.click()
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="tab2"]/div/div[1]/div[2]/div')))  #일반 결제 클릭
        el.click()
        el = wait.until(EC.presence_of_element_located((By.ID, 'btnConfirm')))  #확인 클릭
        el.click()

        #카드번호 입력
        cardno1 = wait.until(EC.presence_of_element_located((By.ID, 'cardno1'))) 
        cardno1.send_keys(0000)
        cardno2 = wait.until(EC.presence_of_element_located((By.ID, 'cardno2')))
        cardno2.send_keys(0000)
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@aria-label="0"]')))
        el.click()
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@aria-label="0"]')))
        el.click()
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@aria-label="0"]')))
        el.click()
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@aria-label="0"]')))
        el.click()

        sleep(1)
        el = wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@id="mtk_cardno4"]/div[@class="dv_transkey_div_2 dv_transkey_div2_Height"]/div[@aria-label="0"]')))
        el.click()
        el = wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@id="mtk_cardno4"]/div[@class="dv_transkey_div_2 dv_transkey_div2_Height"]/div[@aria-label="0"]')))
        el.click()
        el = wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@id="mtk_cardno4"]/div[@class="dv_transkey_div_2 dv_transkey_div2_Height"]/div[@aria-label="0"]')))
        el.click()
        el = wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@id="mtk_cardno4"]/div[@class="dv_transkey_div_2 dv_transkey_div2_Height"]/div[@aria-label="0"]')))
        el.click()

        #유효기간 입력
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.XPATH, '//select[@name="expiryMonth"]/option[9]')))  #9월
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH, '//select[@name="expiryYear"]/option[3]')))  #2024년
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH, '//input[@id="btnOk"]')))  #확인 클릭
        el.click()

        #cvc입력
        el = wait.until(EC.presence_of_element_located((By.ID, 'txtCVC')))
        el.click()
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@aria-label="0"]')))
        el.click()
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@aria-label="0"]')))
        el.click()
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@aria-label="0"]')))
        el.click()

        #결제 비밀번호 입력
        el = wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@id="mtk_pswd"]/div[@class="dv_transkey_div_2 dv_transkey_div2_Height"]/div[@aria-label="0"]')))
        el.click()
        el = wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@id="mtk_pswd"]/div[@class="dv_transkey_div_2 dv_transkey_div2_Height"]/div[@aria-label="0"]')))
        el.click()
        el = wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@id="mtk_pswd"]/div[@class="dv_transkey_div_2 dv_transkey_div2_Height"]/div[@aria-label="0"]')))
        el.click()
        el = wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@id="mtk_pswd"]/div[@class="dv_transkey_div_2 dv_transkey_div2_Height"]/div[@aria-label="0"]')))
        el.click()
        el = wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@id="mtk_pswd"]/div[@class="dv_transkey_div_2 dv_transkey_div2_Height"]/div[@aria-label="0"]')))
        el.click()
        el = wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@id="mtk_pswd"]/div[@class="dv_transkey_div_2 dv_transkey_div2_Height"]/div[@aria-label="0"]')))
        el.click()
        el = wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@id="mtk_pswd"]/div[@class="dv_transkey_div_2 dv_transkey_div2_Height"]/div[@aria-label="0"]')))
        el.click()
        el = wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@id="mtk_pswd"]/div[@class="dv_transkey_div_2 dv_transkey_div2_Height"]/div[@aria-label="0"]')))
        el.click()
        el = wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@id="mtk_pswd"]/div[5]/div[3]')))
        el.click()

        #결제 완료
        el = wait.until(EC.presence_of_element_located((By.ID, 'btnOk')))
        el.click()

        sleep(20)

    #=====>결제화면 캡쳐
        #애터미 결제페이지로 이동
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[1])

        sleep(1)
        self.driver.save_screenshot('KR_CreditCard(APP).png')

    #=====>결제정보 저장
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="body"]/section[2]/table/tbody/tr[1]/td')))   #주문번호 가져오기
        Order_Info['Num'] = el.text
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="body"]/section[2]/table/tbody/tr[4]/td/strong')))    #결제금액 가져오기
        Order_Info['Price'] = el.text

    #=====>주문확인 및 주문취소
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnOrders"]')))  #주문확인 클릭
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'(//button[@name="btnCancelOrder"])[1]')))  #주문취소 클릭
        el.click()
        
        #=====>네이티브앱으로 전환
        self.driver.switch_to.context('NATIVE_APP')
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))   #주문취소 alert 처리
        el.click()
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))   #주문취소 alert 처리
        el.click()

        #=====>웹뷰로 전환
        self.driver.switch_to.context('WEBVIEW_m.com.atomy')

        #주문취소 확인
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//ul[@id="list"]/li[1]/div[1]/span[1]')))
        Order_Info['Cancel'] = el.text

    #=====>결제정보 엑셀에 저장
        wb = openpyxl.load_workbook("APP_CreditCard.xlsx")
        sheet=wb.active
        rows = sheet.max_row

        list_order = list(Order_Info.values())
        sheet.append(list_order)

        if Order_Info['Cancel'] == '주문취소':    
            sheet.cell(rows+1, 5, '취소확인')
        else:
            sheet.cell(rows+1, 5, '취소실패')

        wb.save("APP_CreditCard.xlsx")

    #테스트 종료 후 수행해야하는 코드
    def tearDown(self):
        self.driver.quit()

if __name__ == '__main__':
    reportFoler = "ReportTest_Card"
    unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output=reportFoler))