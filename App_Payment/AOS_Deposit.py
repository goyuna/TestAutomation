#실행 : python AOS_KR.py
import base64
import unittest
import HtmlTestRunner
import os
import time
import openpyxl   
from appium import webdriver
from time import sleep
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException

class DepositTest(unittest.TestCase):
    
    #테스트 실행 전 수행해야하는 코드
    @classmethod
    def setUpClass(cls):

        #Set up Directory & Workbook
        tm = time.localtime()
        str_tm = time.strftime('%Y-%m-%d %H:%M:%S', tm)

        wb=openpyxl.Workbook()
        sheet=wb.active
        sheet['A1'] = str_tm
        sheet.append(['법인명', '계정명', '주문 금액', '주문 번호', '취소 확인', '비고'])
        wb.save('APP_Deposit.xlsx')

    #각 TC 실행 전 수행되는 코드
    def setUp(self):
        app = os.path.join(os.path.dirname(__file__), 'C:\\Users\\고유나\\Desktop\\','m.com.atomy_50_apps.evozi.com.apk')
        app = os.path.abspath(app)

        #Set up appium
        self.driver = webdriver.Remote(
            command_executor='http://127.0.0.1:4723/wd/hub',
            desired_capabilities={
                'app': app,
                'platformName': 'Android',
                'platformVersion': '9.0',
                'deviceName': 'GalaxyS8',
                'automationName': 'uiautomator2',
                'appPackage': 'm.com.atomy',
                'appActivity': 'com.atomy.android.app.views.activities.splash.SplashActivity',
                'udid': 'ce021712d2a8d11e05',
                'chromeOptions' : {'w3c' : False},
                #'noReset': True
            })

    #실행할 테스트 케이스
    def test_KR(self):
        KRID = '120000'
        PW = 'atomy@8580'
        Order_Info = {'Nation':'KR', 'ID': KRID, 'Price': '', 'Num': '','Cancel':'', 'Etc': ''}

        #*************앱 실행 및 권한 허용***********

        # appium의 webdriver를 초기화
        driver = self.driver

        # selenium의 webdriverwait을 사용. element가 나올 때 까지 최대 20초까지 wait
        wait = WebDriverWait(driver, 20)

        #화면 녹화 시작
        #self.driver.start_recording_screen()    

        # 권한 허용 버튼 클릭
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.ID,'com.android.packageinstaller:id/permission_allow_button')))
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.ID,'com.android.packageinstaller:id/permission_allow_button')))
        el.click()
        sleep(30)

        #=====>웹뷰로 전환
        self.driver.switch_to.context('WEBVIEW_m.com.atomy')

        #공지사항 하루종일 보지않기
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//section[@id="popup"]/div/div[2]/div/span[1]')))
        el.click()
        sleep(2)

    #=====>로그인
        #로그인버튼 클릭
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//header[@id="header"]/div[2]/span[1]/a')))
        el.click()
        sleep(2)

        #로그인창으로 Switch
        window_handles = self.driver.window_handles
        #print(window_handles)
        self.driver.switch_to.window(window_handles[1])

        #로그인 진행
        id = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@id="UserId"]')))
        id.send_keys(KRID)
        pw = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@id="Password"]')))
        pw.send_keys(PW)
        sleep(1)
        self.driver.find_element_by_xpath('//button[@id="btnLogin"]').click()
        sleep(3)

    #=====>쇼핑몰로 이동
        self.driver.switch_to.window(window_handles[0])

        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//nav[@id="nav"]/a[1]')))
        el.click()
        sleep(5)

        #쇼핑몰창으로 Switch
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[2])

        #튜토리얼 닫기
        self.driver.find_element_by_xpath('//section[@id="toutorialDialog"]/div/button').click()
        sleep(2)

    #=====>최저가격 상품 찾기
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//select[@name="orderType"]/option[3]')))     #낮은 가격순
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//ul[@id="list"]/li[2]/div/div/div/span/a')))
        el.click()
        sleep(2)

        #바로구매 클릭
        window_handles = self.driver.window_handles
        #print(window_handles)
        self.driver.switch_to.window(window_handles[3])
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnBuy"]')))
        el.click() 
        sleep(1)
        el.click()
        sleep(1)

    #=====>배송 정보 입력
        sleep(3)
        self.driver.switch_to.window(window_handles[2])
        self.driver.execute_script('window.scrollTo(0, 600)')
        sleep(1)                                     
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//section[@id="tabInsertShippingInfo"]')))
        el.click() 
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//div[@id="radioShippingInfo"]')))   #주문자와 동일
        el.click()

        #결제 방법 입력
        self.driver.execute_script('window.scrollTo(0, 2400)') 
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//form[@id="frmBuy"]/ul/li[4]/section/div[1]/div[2]/label'))) #무통장입금 선택
        el.click() 
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//select[@name="accountPayment.bankAccount.bank.code"]/option[2]')))  #기업은행 선택
        el.click() 
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@name="paymentItemVAccountMobile2"]')))  #전화번호 입력
        el.send_keys('111')
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@name="paymentItemVAccountMobile3"]')))  #전화번호 입력
        el.send_keys('1111')
        self.driver.hide_keyboard()

    #=====>결제 하기
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        sleep(1)
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//input[@id="checkBuyAgree"]')))  #동의하기 선택
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnBuy"]')))  #결제하기 클릭
        el.click() 
        
    #=====>결제화면 캡쳐
        sleep(1)
        self.driver.save_screenshot('KR_Deposit(APP).png')

    #=====>결제정보 저장
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="body"]/section[2]/table/tbody/tr[1]/td')))   #주문번호 가져오기
        Order_Info['Num'] = el.text
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="body"]/section[2]/table/tbody/tr[4]/td/strong')))    #결제금액 가져오기
        Order_Info['Price'] = el.text

    #=====>주문확인 및 주문취소
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnOrders"]')))  #주문확인 클릭
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'(//button[@name="btnCancelOrder"])[1]')))  #주문취소 클릭
        el.click()
        
        #=====>네이티브앱으로 전환
        self.driver.switch_to.context('NATIVE_APP')
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))   #주문취소 alert 처리
        el.click()
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))   #주문취소 alert 처리
        el.click()

        #=====>웹뷰로 전환
        self.driver.switch_to.context('WEBVIEW_m.com.atomy')

        #주문취소 확인
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//ul[@id="list"]/li[1]/div[1]/span[1]')))
        Order_Info['Cancel'] = el.text

    #=====>결제정보 엑셀에 저장
        wb = openpyxl.load_workbook("APP_Deposit.xlsx")
        sheet=wb.active
        rows = sheet.max_row

        list_order = list(Order_Info.values())
        sheet.append(list_order)

        if Order_Info['Cancel'] == '주문취소':    
            sheet.cell(rows+1, 5, '취소확인')
        else:
            sheet.cell(rows+1, 5, '취소실패')

        wb.save("APP_Deposit.xlsx")

    def test_KR_SMS(self):
        KRID = '120000'
        PW = 'atomy@8580'
        Order_Info = {'Nation':'KR', 'ID': KRID, 'Price': '', 'Num': '','Cancel':'', 'Etc': 'SMS결제'}

    #=====>앱 실행 및 권한 허용
        driver = self.driver
        wait = WebDriverWait(driver, 20)

        # 권한 허용 버튼 클릭
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.ID,'com.android.packageinstaller:id/permission_allow_button')))
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.ID,'com.android.packageinstaller:id/permission_allow_button')))
        el.click()
        sleep(30)

        #=====>웹뷰로 전환
        self.driver.switch_to.context('WEBVIEW_m.com.atomy')

        #공지사항 하루종일 보지않기
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//section[@id="popup"]/div/div[2]/div/span[1]')))
        el.click()
        sleep(2)

    #=====>로그인
        #로그인버튼 클릭
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//header[@id="header"]/div[2]/span[1]/a')))
        el.click()
        sleep(2)

        #로그인창으로 Switch
        window_handles = self.driver.window_handles
        #print(window_handles)
        self.driver.switch_to.window(window_handles[1])

        #로그인 진행
        id = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@id="UserId"]')))
        id.send_keys(KRID)
        pw = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@id="Password"]')))
        pw.send_keys(PW)
        sleep(1)
        self.driver.find_element_by_xpath('//button[@id="btnLogin"]').click()
        sleep(1)

    #=====>쇼핑몰로 이동
        self.driver.switch_to.window(window_handles[0])

        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//nav[@id="nav"]/a[1]')))
        el.click()
        sleep(5)

        #쇼핑몰창으로 Switch
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[2])

        #튜토리얼 닫기
        self.driver.find_element_by_xpath('//section[@id="toutorialDialog"]/div/button').click()
        sleep(2)

    #=====>최저가격 상품 찾기
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//select[@name="orderType"]/option[3]')))     #낮은 가격순
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//ul[@id="list"]/li[2]/div/div/div/span/a')))
        el.click()
        sleep(2)

        #바로구매 클릭
        window_handles = self.driver.window_handles
        #print(window_handles)
        self.driver.switch_to.window(window_handles[3])
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnBuy"]')))
        el.click() 
        sleep(1)
        el.click()
        sleep(1)

    #=====>배송 정보 입력
        sleep(3)
        self.driver.switch_to.window(window_handles[0])
        self.driver.execute_script('window.scrollTo(0, 600)')
        sleep(1)                                     
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//section[@id="tabInsertShippingInfo"]')))
        el.click() 
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//div[@id="radioShippingInfo"]')))   #주문자와 동일
        el.click()

        #결제 방법 입력
        self.driver.execute_script('window.scrollTo(0, 2500)') 
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//form[@id="frmBuy"]/ul/li[4]/section/div[1]/div[3]/label'))) #SMS결제 선택
        el.click() 
        el = wait.until(EC.element_to_be_clickable((By.ID,'paymentItemSMSMobile')))  #전화번호 입력
        el.send_keys('0101111111')
        self.driver.hide_keyboard()

    #=====>결제 하기
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        sleep(1)
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//input[@id="checkBuyAgree"]')))  #동의하기 선택
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnBuy"]')))  #결제하기 클릭
        el.click() 
        
    #=====>결제화면 캡쳐
        sleep(1)
        self.driver.save_screenshot('KR_SMS(APP).png')

    #=====>결제정보 저장
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="body"]/section[2]/table/tbody/tr[1]/td')))   #주문번호 가져오기
        Order_Info['Num'] = el.text
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="body"]/section[2]/table/tbody/tr[4]/td/strong')))    #결제금액 가져오기
        Order_Info['Price'] = el.text

    #=====>주문확인 및 주문취소
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnOrders"]')))  #주문확인 클릭
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'(//button[@name="btnCancelOrder"])[1]')))  #주문취소 클릭
        el.click()
        
        #=====>네이티브앱으로 전환
        self.driver.switch_to.context('NATIVE_APP')
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))   #주문취소 alert 처리
        el.click()
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))   #주문취소 alert 처리
        el.click()

        #=====>웹뷰로 전환
        self.driver.switch_to.context('WEBVIEW_m.com.atomy')

        #주문취소 확인
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//ul[@id="list"]/li[1]/div[1]/span[1]')))
        Order_Info['Cancel'] = el.text

    #=====>결제정보 엑셀에 저장
        wb = openpyxl.load_workbook("APP_Deposit.xlsx")
        sheet=wb.active
        rows = sheet.max_row

        list_order = list(Order_Info.values())
        sheet.append(list_order)

        if Order_Info['Cancel'] == '주문취소':    
            sheet.cell(rows+1, 5, '취소확인')
        else:
            sheet.cell(rows+1, 5, '취소실패')

        wb.save("APP_Deposit.xlsx")

    def test_US(self):
        USID = '27057016'
        PW = 'atomy@8580'
        Order_Info = {'Nation':'US', 'ID': USID, 'Price': '', 'Num': '','Cancel':'', 'Etc': ''}

        driver = self.driver

        wait = WebDriverWait(driver, 20)

        # 권한 허용 버튼 클릭
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.ID,'com.android.packageinstaller:id/permission_allow_button')))
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.ID,'com.android.packageinstaller:id/permission_allow_button')))
        el.click()
        sleep(30)

        #=====>웹뷰로 전환
        self.driver.switch_to.context('WEBVIEW_m.com.atomy')

        #공지사항 하루종일 보지않기
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//section[@id="popup"]/div/div[2]/div/span[1]')))
        el.click()
        sleep(2)

    #=====>법인 변경
        #LNB버튼 클릭
        menu = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnMenu"]')))
        menu.click()

        #설정버튼 클릭
        sleep(1)
        setting = wait.until(EC.element_to_be_clickable((By.XPATH,'//nav[@id="gnb"]/div[1]/span/a')))
        setting.click()
        sleep(3)

        #=====>네이티브앱으로 전환
        #설정으로 Switch
        self.driver.switch_to.context('NATIVE_APP')

        #국가 선택
        el = wait.until(EC.element_to_be_clickable((By.ID,'m.com.atomy:id/spinnerRegion')))
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//android.widget.TextView[@text="미국"]')))
        el.click()

        #앱 재시작
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))
        el.click()

        sleep(10)

    #=====>언어 변경
        #=====>웹뷰로 전환
        self.driver.switch_to.context('WEBVIEW_m.com.atomy')

        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[0])

        #공지사항 하루종일 보지않기
        try:
            el = self.driver.find_element_by_xpath('//section[@id="popup"]/div/div[2]/div/span[1]')
        except NoSuchElementException:
            pass
        else:
            el.click()

        #LNB버튼 클릭
        sleep(2)
        menu = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnMenu"]')))
        menu.click()

        #설정버튼 클릭
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//nav[@id="gnb"]/div[1]/span/a')))
        el.click()
        sleep(5)

        #=====>네이티브앱으로 전환
        #설정으로 Switch
        self.driver.switch_to.context('NATIVE_APP')

        #언어 선택
        el = wait.until(EC.element_to_be_clickable((By.ID,'m.com.atomy:id/spinnerLanguage')))
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//android.widget.TextView[@text="English (United States)"]')))
        el.click()

        #앱 재시작
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))
        el.click()

        sleep(10)

    #=====>로그인
        #=====>웹뷰로 전환
        self.driver.switch_to.context('WEBVIEW_m.com.atomy')

        #공지사항 하루종일 보지않기
        try:
            el = self.driver.find_element_by_id('btnCloseWithSetCookie')
        except NoSuchElementException:
            pass
        else:
            el.click()

        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//header[@id="header"]/div[2]/span[1]/a')))
        el.click()

        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[1])

        id = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@name="UserId"]')))
        id.send_keys(USID)
        pw = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@name="Password"]')))
        #pw.click()
        pw.send_keys(PW)
        sleep(1)
        self.driver.find_element_by_xpath('//button[@id="btnLogin"]').click()
        sleep(2)

    #=====>쇼핑몰로 이동
        self.driver.switch_to.window(window_handles[0])

        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//nav[@id="nav"]/a[1]')))
        el.click()
        sleep(3)

        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[1])

        #튜토리얼 닫기
        self.driver.find_element_by_xpath('//section[@id="toutorialDialog"]/div/button').click()
        sleep(2)

    #=====>최저가격 상품 찾기
        el = wait.until(EC.element_to_be_clickable((By.XPATH, '//button[@id="btnActiveSearcher"]')))
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@id="autocomplete"]')))
        el.send_keys('shopping bag')
        el = wait.until(EC.element_to_be_clickable((By.XPATH, '//button[@id="btnSearch"]')))
        el.click()
        sleep(3)

        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[2])

        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//ul[@id="list"]/li[1]/div/div/div/span/a')))
        el.click() 
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnBuy"]')))
        el.click() 
        sleep(1)
        el.click()
        sleep(1)

    #=====>배송 정보 입력
        sleep(3)
        self.driver.switch_to.window(window_handles[1])
        self.driver.execute_script('window.scrollTo(0, 600)')
        sleep(1)                                     
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//section[@id="tabInsertShippingInfo"]/div/span[2]/label'))) #주문자와 동일
        el.click() 

        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@name="shipping.recipientFirstName"]')))   #First name
        el.send_keys('atomy')
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@name="shipping.recipientLastName"]')))   #Last name
        el.send_keys('atomy')
        self.driver.hide_keyboard()

        #결제 방법 입력
        self.driver.execute_script('window.scrollTo(0, 1800)') 
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//ul[@id="paymentTypeList"]/li[2]'))) #Bank Deposit 선택
        el.click() 

    #=====>결제 하기
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        sleep(1)
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//input[@name="checkBuyAgree"]')))  #동의하기 선택
        el.click()
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//button[@id="btnCalculate"]')))  #confirm order 선택
        el.click()
        sleep(2)

        #tax detail 창으로 전환
        self.driver.switch_to.window(window_handles[1]) 
        el = self.driver.find_element_by_xpath('//section[@id="dialogSalesTax"]/div/button')
        self.driver.execute_script("arguments[0].click();", el)     #tax detail 창 닫기

        sleep(2)
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnBuy"]')))  #결제하기 클릭
        el.click() 
        
    #=====>결제화면 캡쳐
        sleep(2)
        self.driver.save_screenshot('US_Deposit(APP).png')

    #=====>결제정보 저장
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="body"]/section[2]/table/tbody/tr[1]/td')))   #주문번호 가져오기
        Order_Info['Num'] = el.text
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="body"]/section[2]/table/tbody/tr[3]/td/strong')))    #결제금액 가져오기
        Order_Info['Price'] = el.text

    #=====>주문확인 및 주문취소
        self.driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnOrders"]')))  #주문확인 클릭
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'(//button[@name="btnCancelOrder"])')))  #주문취소 클릭
        el.click()
        
        #=====>네이티브앱으로 전환
        self.driver.switch_to.context('NATIVE_APP')
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))   #주문취소 alert 처리
        el.click()
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))   #주문취소 alert 처리
        el.click()
        sleep(1)

        #=====>웹뷰로 전환
        self.driver.switch_to.context('WEBVIEW_m.com.atomy')

        #주문취소 확인
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//ul[@id="list"]/li/div[1]/span[1]')))
        Order_Info['Cancel'] = el.text

    #=====>결제정보 엑셀에 저장
        wb = openpyxl.load_workbook("APP_Deposit.xlsx")
        sheet=wb.active
        rows = sheet.max_row

        list_order = list(Order_Info.values())
        sheet.append(list_order)

        if Order_Info['Cancel'] == 'Cancelled Order':    
            sheet.cell(rows+1, 5, '취소확인')
        else:
            sheet.cell(rows+1, 5, '취소실패')

        wb.save("APP_Deposit.xlsx")

    def test_ID(self):
        IDID = '20852325'
        PW = 'atomy@8580'
        Order_Info = {'Nation': 'ID', 'ID': IDID, 'Price':'', 'Num': '', 'Cancel':'', 'Etc':''}

        driver = self.driver

        wait = WebDriverWait(driver, 20)

        # 권한 허용 버튼 클릭
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.ID,'com.android.packageinstaller:id/permission_allow_button')))
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.ID,'com.android.packageinstaller:id/permission_allow_button')))
        el.click()
        sleep(30)

        #=====>웹뷰로 전환
        self.driver.switch_to.context('WEBVIEW_m.com.atomy')

        #공지사항 하루종일 보지않기
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//section[@id="popup"]/div/div[2]/div/span[1]')))
        el.click()
        sleep(2)

    #=====>법인 변경
        #LNB버튼 클릭
        menu = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnMenu"]')))
        menu.click()

        #설정버튼 클릭
        sleep(1)
        setting = wait.until(EC.element_to_be_clickable((By.XPATH,'//nav[@id="gnb"]/div[1]/span/a')))
        setting.click()
        sleep(5)

        #=====>네이티브앱으로 전환
        #설정으로 Switch
        self.driver.switch_to.context('NATIVE_APP')

        #국가 선택
        el = wait.until(EC.element_to_be_clickable((By.ID,'m.com.atomy:id/spinnerRegion')))
        el.click()
        sleep(1)

        #spinner scroll down 
        el_tap = driver.find_element_by_xpath('//android.widget.TextView[@text="멕시코"]')
        el_dragto = driver.find_element_by_xpath('//android.widget.TextView[@text="싱가포르"]')
        driver.scroll(el_tap, el_dragto)    # : 멕시코를 싱가포르까지 scroll
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//android.widget.TextView[@text="인도네시아"]')))
        el.click()

        #앱 재시작
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))
        el.click()

        sleep(15)

    #=====>언어 변경
        #=====>웹뷰로 전환
        self.driver.switch_to.context('WEBVIEW_m.com.atomy')

        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[0])

        #공지사항 하루종일 보지않기
        try:
            el = self.driver.find_element_by_id('btnCloseWithSetCookie')
        except NoSuchElementException:
            pass
        else:
            el.click()

        #LNB버튼 클릭
        sleep(2)
        menu = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnMenu"]')))
        menu.click()

        #설정버튼 클릭
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//nav[@id="gnb"]/div[1]/span/a')))
        el.click()
        sleep(5)

        #=====>네이티브앱으로 전환
        #설정으로 Switch
        self.driver.switch_to.context('NATIVE_APP')

        #언어 선택
        el = wait.until(EC.element_to_be_clickable((By.ID,'m.com.atomy:id/spinnerLanguage')))
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//android.widget.TextView[@text="Bahasa Indonesia (INDONESIA)"]')))
        el.click()

        #앱 재시작
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))
        el.click()

        sleep(10)

    #=====>로그인
        #=====>웹뷰로 전환
        self.driver.switch_to.context('WEBVIEW_m.com.atomy')

        #공지사항 하루종일 보지않기
        try:
            el = self.driver.find_element_by_id('btnCloseWithSetCookie')
        except NoSuchElementException:
            pass
        else:
            el.click()

        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//header[@id="header"]/div[2]/span[1]/a')))
        el.click()

        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[1])

        id = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@name="UserId"]')))
        id.send_keys(IDID)
        pw = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@name="Password"]')))
        pw.send_keys(PW)
        sleep(1)
        self.driver.find_element_by_xpath('//button[@id="btnLogin"]').click()
        sleep(2)

    #=====>쇼핑몰로 이동
        self.driver.switch_to.window(window_handles[0])

        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//nav[@id="nav"]/a[1]')))
        el.click()
        sleep(3)

        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[1])
        sleep(1)

        #튜토리얼 닫기
        self.driver.find_element_by_xpath('//section[@id="toutorialDialog"]/div/button').click()
        sleep(2)

    #=====>최저가격 상품 찾기
        el = wait.until(EC.element_to_be_clickable((By.ID, 'btnActiveSearcher')))
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@id="autocomplete"]')))
        el.send_keys('shopping bag')
        el = wait.until(EC.element_to_be_clickable((By.XPATH, '//button[@id="btnSearch"]')))
        el.click()
        sleep(3)

        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[2])

        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//ul[@id="list"]/li[1]/div/div/div/span/a')))
        el.click() 
        sleep(1)

        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnBuy"]')))
        el.click() 

        self.driver.switch_to.context('NATIVE_APP')
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))   #alert 처리
        el.click()

        self.driver.switch_to.context('WEBVIEW_m.com.atomy')
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnBuy"]')))
        el.click()
        sleep(1)

    #=====>배송 정보 입력
        sleep(3)
        self.driver.switch_to.window(window_handles[1])
        self.driver.execute_script('window.scrollTo(0, 600)')
        sleep(1)                                     
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//div[@id="shippingInfoSelect"]/span[1]'))) #주문자와 동일
        el.click() 
   
        #결제 방법 입력
        self.driver.execute_script('window.scrollTo(0, 1400)') 
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//ul[@id="paymentTypeList"]/li[2]'))) #Virtual Account 선택
        el.click() 
   
    #=====>결제 하기
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        sleep(1)
        el = wait.until(EC.presence_of_element_located((By.ID, 'checkBuyAgree')))  #동의하기 선택
        el.click()
        el = wait.until(EC.presence_of_element_located((By.ID, 'btnCalculate')))  #confirm order 선택
        el.click()
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.ID,'btnBuy')))  #결제하기 클릭
        el.click() 
        sleep(10)

        #은행사 화면으로 전환
        window_handles = self.driver.window_handles
        print(window_handles)
        self.driver.switch_to.window(window_handles[3])
        self.driver.execute_script('document.querySelector("#optionsCENA").click();') #은행 선택
        self.driver.execute_script('document.getElementById("submit_button").click();') #결제하기 클릭
        sleep(3)

    #=====>결제화면 캡쳐
        sleep(2)

        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[1])
        self.driver.save_screenshot('ID_Deposit(APP).png')

    #=====>결제정보 저장
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="body"]/section[2]/table/tbody/tr[1]/td')))   #주문번호 가져오기
        Order_Info['Num'] = el.text
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="body"]/section[2]/table/tbody/tr[3]/td/strong')))    #결제금액 가져오기
        Order_Info['Price'] = el.text

    #=====>주문확인 및 주문취소
        self.driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'//button[@id="btnOrders"]')))  #주문확인 클릭
        el.click()
        el = wait.until(EC.element_to_be_clickable((By.XPATH,'(//button[@name="btnCancelOrder"])')))  #주문취소 클릭
        el.click()
        sleep(1)
        
        #=====>네이티브앱으로 전환
        self.driver.switch_to.context('NATIVE_APP')
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))   #주문취소 alert 처리
        el.click()
        sleep(1)
        el = wait.until(EC.element_to_be_clickable((By.ID,'android:id/button1')))   #주문취소 alert 처리
        el.click()
        sleep(1)

        #=====>웹뷰로 전환
        self.driver.switch_to.context('WEBVIEW_m.com.atomy')

        #주문취소 확인
        el = wait.until(EC.presence_of_element_located((By.XPATH, '//ul[@id="list"]/li/div[1]/span[1]')))
        Order_Info['Cancel'] = el.text

    #=====>결제정보 엑셀에 저장
        wb = openpyxl.load_workbook("APP_Deposit.xlsx")
        sheet=wb.active
        rows = sheet.max_row

        list_order = list(Order_Info.values())
        sheet.append(list_order)

        if Order_Info['Cancel'] == 'Pesanan Dibatalkan':    
            sheet.cell(rows+1, 5, '취소확인')
        else:
            sheet.cell(rows+1, 5, '취소실패')

        wb.save("APP_Deposit.xlsx")
        
    #각 TC 종료 후 수행되는 코드
    def tearDown(self):
        self.driver.quit()

#unittest 실행
if __name__ == '__main__':
    reportFoler = "ReportTest_Deposit"
    unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output=reportFoler))
